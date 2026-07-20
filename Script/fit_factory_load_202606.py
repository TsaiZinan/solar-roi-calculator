from __future__ import annotations

import csv
import glob
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/cai/SynologyDrive/Project/#ProjectWork-000000-光伏收益计算")
ORDER_FILE = sorted(glob.glob(str(ROOT / "数据" / "TEMP" / "*.xlsx")))[0]
DAILY_CSV_GLOB = str(ROOT / "数据" / "202606*" / "*.csv")


def parse_dt(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def to_float(value):
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def load_site_total_load():
    rows = []
    for path in sorted(glob.glob(DAILY_CSV_GLOB)):
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                ts = parse_dt(row.get("时间"))
                if not ts:
                    continue
                pv_kw = max(0.0, to_float(row.get("光伏发电功率(kW)")))
                raw_load_kw = to_float(row.get("负载功率(kW)"))
                total_load_kw = max(0.0, pv_kw - raw_load_kw)
                rows.append((ts, total_load_kw))
    rows.sort(key=lambda item: item[0])
    return rows


def load_order_power_by_timestamp(valid_timestamps):
    valid_set = set(valid_timestamps)
    power_by_ts = defaultdict(float)

    wb = load_workbook(ORDER_FILE, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {name: i for i, name in enumerate(header)}

    valid_orders = 0
    for row in rows:
        status = row[idx["充电状态"]]
        pay_status = row[idx["支付状态"]]
        abnormal = row[idx["是否异常"]]
        if not (status == "已完成" and pay_status == "已支付" and abnormal in ("正常", None, "")):
            continue

        start = parse_dt(row[idx["开始充电时间"]])
        end = parse_dt(row[idx["充电完成时间"]])
        kwh = to_float(row[idx["充电量（度）"]])
        if not start or not end or end <= start or kwh <= 0:
            continue

        valid_orders += 1
        total_seconds = (end - start).total_seconds()
        if total_seconds <= 0:
            continue

        cursor = start
        while cursor < end:
            slot_start = cursor.replace(
                minute=(cursor.minute // 5) * 5,
                second=0,
                microsecond=0,
            )
            slot_end = slot_start + timedelta(minutes=5)
            segment_end = min(end, slot_end)
            overlap_seconds = (segment_end - cursor).total_seconds()
            if overlap_seconds > 0 and slot_start in valid_set:
                energy_kwh = kwh * overlap_seconds / total_seconds
                power_kw = energy_kwh / (5.0 / 60.0)
                power_by_ts[slot_start] += power_kw
            cursor = segment_end

    return power_by_ts, valid_orders


def aggregate(rows):
    by_day = defaultdict(lambda: {"factory_kwh": 0.0, "site_kwh": 0.0, "order_kwh": 0.0})
    by_day_hour = defaultdict(lambda: {"factory_kwh": 0.0, "site_kwh": 0.0, "order_kwh": 0.0, "samples": 0})
    by_hour = defaultdict(lambda: {"factory_kwh": 0.0, "site_kwh": 0.0, "order_kwh": 0.0})

    detail_rows = []
    for ts, total_load_kw, order_power_kw, factory_power_kw in rows:
        date_str = ts.strftime("%Y%m%d")
        hour = ts.hour
        site_kwh = total_load_kw * (5.0 / 60.0)
        order_kwh = order_power_kw * (5.0 / 60.0)
        factory_kwh = factory_power_kw * (5.0 / 60.0)

        by_day[date_str]["factory_kwh"] += factory_kwh
        by_day[date_str]["site_kwh"] += site_kwh
        by_day[date_str]["order_kwh"] += order_kwh

        by_day_hour[(date_str, hour)]["factory_kwh"] += factory_kwh
        by_day_hour[(date_str, hour)]["site_kwh"] += site_kwh
        by_day_hour[(date_str, hour)]["order_kwh"] += order_kwh
        by_day_hour[(date_str, hour)]["samples"] += 1

        by_hour[hour]["factory_kwh"] += factory_kwh
        by_hour[hour]["site_kwh"] += site_kwh
        by_hour[hour]["order_kwh"] += order_kwh

        detail_rows.append(
            [
                ts.strftime("%Y-%m-%d %H:%M:%S"),
                round(total_load_kw, 4),
                round(order_power_kw, 4),
                round(factory_power_kw, 4),
            ]
        )

    return by_day, by_day_hour, by_hour, detail_rows


def main():
    site_rows = load_site_total_load()
    timestamps = [ts for ts, _ in site_rows]
    order_power_by_ts, valid_orders = load_order_power_by_timestamp(timestamps)

    fitted_rows = []
    clipped_count = 0
    for ts, total_load_kw in site_rows:
        order_power_kw = max(0.0, order_power_by_ts.get(ts, 0.0))
        factory_power_kw = total_load_kw - order_power_kw
        if factory_power_kw < 0:
            clipped_count += 1
            factory_power_kw = 0.0
        fitted_rows.append((ts, total_load_kw, order_power_kw, factory_power_kw))

    by_day, by_day_hour, by_hour, detail_rows = aggregate(fitted_rows)

    out_dir = ROOT / "报告" / "json"
    out_dir.mkdir(parents=True, exist_ok=True)
    detail_csv = out_dir / "工厂用电拟合_202606_5分钟明细.csv"
    day_hour_csv = out_dir / "工厂用电拟合_202606_逐日逐小时.csv"
    day_csv = out_dir / "工厂用电拟合_202606_逐日汇总.csv"
    hour_csv = out_dir / "工厂用电拟合_202606_逐小时汇总.csv"
    report_md = ROOT / "报告" / "6月工厂用电拟合分析.md"

    with open(detail_csv, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["时间", "总站总负载(kW)", "订单折算充电功率(kW)", "拟合工厂功率(kW)"])
        writer.writerows(detail_rows)

    with open(day_hour_csv, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["日期", "小时", "总站总负载(度)", "订单充电量(度)", "拟合工厂用电(度)", "拟合工厂平均功率(kW)"])
        for date_str, hour in sorted(by_day_hour):
            data = by_day_hour[(date_str, hour)]
            samples = max(1, data["samples"])
            avg_factory_kw = data["factory_kwh"] / (samples * (5.0 / 60.0))
            writer.writerow(
                [
                    date_str,
                    f"{hour:02d}:00",
                    round(data["site_kwh"], 4),
                    round(data["order_kwh"], 4),
                    round(data["factory_kwh"], 4),
                    round(avg_factory_kw, 4),
                ]
            )

    with open(day_csv, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["日期", "总站总负载(度)", "订单充电量(度)", "拟合工厂用电(度)", "工厂占总站负载比"])
        for date_str in sorted(by_day):
            data = by_day[date_str]
            ratio = data["factory_kwh"] / data["site_kwh"] if data["site_kwh"] else ""
            writer.writerow(
                [
                    date_str,
                    round(data["site_kwh"], 4),
                    round(data["order_kwh"], 4),
                    round(data["factory_kwh"], 4),
                    round(ratio, 6) if ratio != "" else "",
                ]
            )

    with open(hour_csv, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["小时", "总站总负载(度)", "订单充电量(度)", "拟合工厂用电(度)", "工厂占总站负载比"])
        for hour in range(24):
            data = by_hour[hour]
            ratio = data["factory_kwh"] / data["site_kwh"] if data["site_kwh"] else ""
            writer.writerow(
                [
                    f"{hour:02d}:00",
                    round(data["site_kwh"], 4),
                    round(data["order_kwh"], 4),
                    round(data["factory_kwh"], 4),
                    round(ratio, 6) if ratio != "" else "",
                ]
            )

    total_factory_kwh = sum(data["factory_kwh"] for data in by_day.values())
    total_site_kwh = sum(data["site_kwh"] for data in by_day.values())
    total_order_kwh = sum(data["order_kwh"] for data in by_day.values())
    top_days = sorted(by_day.items(), key=lambda kv: kv[1]["factory_kwh"], reverse=True)[:10]
    top_hours = sorted(by_hour.items(), key=lambda kv: kv[1]["factory_kwh"], reverse=True)

    lines = []
    lines.append("# 6月工厂用电拟合分析")
    lines.append("")
    lines.append("- 方法：`工厂功率 = 总站总负载 - 订单折算充电功率`，结果小于 0 时按 0 截断。")
    lines.append("- 假设：单笔订单在其开始-结束时段内功率近似均匀。")
    lines.append(f"- 订单文件：`{Path(ORDER_FILE).name}`")
    lines.append(f"- 有效订单数：`{valid_orders}`")
    lines.append(f"- 出现负值并被截断的 5 分钟点数：`{clipped_count}`")
    lines.append(f"- 6 月总站总负载：`{total_site_kwh:.2f}` 度")
    lines.append(f"- 6 月订单折算充电量：`{total_order_kwh:.2f}` 度")
    lines.append(f"- 6 月拟合工厂用电：`{total_factory_kwh:.2f}` 度")
    lines.append(f"- 工厂占总站总负载比：`{(total_factory_kwh / total_site_kwh * 100) if total_site_kwh else 0:.2f}%`")
    lines.append("")
    lines.append("## 拟合工厂用电最多的日期 Top 10")
    for date_str, data in top_days:
        ratio = data["factory_kwh"] / data["site_kwh"] * 100 if data["site_kwh"] else 0.0
        lines.append(
            f"- `{date_str}`：工厂 `{data['factory_kwh']:.2f}` 度，总站 `{data['site_kwh']:.2f}` 度，占比 `{ratio:.1f}%`"
        )
    lines.append("")
    lines.append("## 按小时汇总的工厂用电曲线")
    for hour, data in top_hours:
        ratio = data["factory_kwh"] / data["site_kwh"] * 100 if data["site_kwh"] else 0.0
        lines.append(
            f"- `{hour:02d}:00`：工厂 `{data['factory_kwh']:.2f}` 度，总站 `{data['site_kwh']:.2f}` 度，占比 `{ratio:.1f}%`"
        )

    report_md.write_text("\n".join(lines), encoding="utf-8")

    print("detail_csv", detail_csv)
    print("day_hour_csv", day_hour_csv)
    print("day_csv", day_csv)
    print("hour_csv", hour_csv)
    print("report_md", report_md)
    print("total_site_kwh", round(total_site_kwh, 2))
    print("total_order_kwh", round(total_order_kwh, 2))
    print("total_factory_kwh", round(total_factory_kwh, 2))
    print("clipped_count", clipped_count)
    print("top_days")
    for date_str, data in top_days[:5]:
        print(date_str, round(data["factory_kwh"], 2), round(data["site_kwh"], 2), round(data["order_kwh"], 2))
    print("top_hours")
    for hour, data in top_hours[:8]:
        print(f"{hour:02d}:00", round(data["factory_kwh"], 2), round(data["site_kwh"], 2), round(data["order_kwh"], 2))


if __name__ == "__main__":
    main()
