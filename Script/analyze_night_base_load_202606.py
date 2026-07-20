from __future__ import annotations

import csv
import glob
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/cai/SynologyDrive/Project/#ProjectWork-000000-光伏收益计算")
ORDER_FILE = sorted(glob.glob(str(ROOT / "数据" / "TEMP" / "*.xlsx")))[0]
DAILY_CSV_GLOB = str(ROOT / "数据" / "202606*" / "*.csv")

PV_IDLE_THRESHOLD_KW = 0.5
EV_IDLE_THRESHOLD_KW = 1.0
ESS_IDLE_THRESHOLD_KW = 3.0
NIGHT_HOURS = set(range(0, 6)) | set(range(18, 24))


def parse_dt(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def to_float(value):
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def load_site_rows():
    rows = []
    for path in sorted(glob.glob(DAILY_CSV_GLOB)):
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                ts = parse_dt(row.get("时间"))
                if not ts:
                    continue
                pv_kw = max(0.0, to_float(row.get("光伏发电功率(kW)")))
                raw_load_kw = to_float(row.get("负载功率(kW)"))
                total_load_kw = max(0.0, pv_kw - raw_load_kw)
                ess_kw = to_float(row.get("储能有功功率(kW)"))
                rows.append(
                    {
                        "ts": ts,
                        "pv_kw": pv_kw,
                        "total_load_kw": total_load_kw,
                        "ess_kw": ess_kw,
                    }
                )
    rows.sort(key=lambda item: item["ts"])
    return rows


def load_order_power_by_timestamp(valid_timestamps):
    valid_set = set(valid_timestamps)
    power_by_ts = defaultdict(float)

    wb = load_workbook(ORDER_FILE, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {name: i for i, name in enumerate(header)}

    for row in rows:
        status = row[idx["充电状态"]]
        pay_status = row[idx["支付状态"]]
        abnormal = row[idx["是否异常"]]
        if not (status == "已完成" and pay_status == "已支付" and abnormal in ("正常", None, "")):
            continue

        start = parse_dt(row[idx["开始充电时间"]])
        end = parse_dt(row[idx["充电完成时间"]])
        kwh = to_float(row[idx["充电量（度）"]])
        if not start or not end or end <= start or kwh <= 0:
            continue

        total_seconds = (end - start).total_seconds()
        cursor = start
        while cursor < end:
            slot_start = cursor.replace(
                minute=(cursor.minute // 5) * 5,
                second=0,
                microsecond=0,
            )
            slot_end = slot_start + timedelta(minutes=5)
            segment_end = min(end, slot_end)
            overlap_seconds = (segment_end - cursor).total_seconds()
            if overlap_seconds > 0 and slot_start in valid_set:
                energy_kwh = kwh * overlap_seconds / total_seconds
                power_kw = energy_kwh / (5.0 / 60.0)
                power_by_ts[slot_start] += power_kw
            cursor = segment_end

    return power_by_ts


def mean(values):
    return sum(values) / len(values) if values else 0.0


def percentile(values, ratio):
    if not values:
        return 0.0
    values = sorted(values)
    if len(values) == 1:
        return values[0]
    pos = (len(values) - 1) * ratio
    left = int(pos)
    right = min(left + 1, len(values) - 1)
    frac = pos - left
    return values[left] * (1 - frac) + values[right] * frac


def main():
    site_rows = load_site_rows()
    timestamps = [row["ts"] for row in site_rows]
    order_power_by_ts = load_order_power_by_timestamp(timestamps)

    night_rows = []
    hourly_groups = defaultdict(list)
    for row in site_rows:
        ts = row["ts"]
        if ts.hour not in NIGHT_HOURS:
            continue
        order_kw = max(0.0, order_power_by_ts.get(ts, 0.0))
        enriched = {
            "ts": ts,
            "pv_kw": row["pv_kw"],
            "total_load_kw": row["total_load_kw"],
            "ess_kw": row["ess_kw"],
            "order_kw": order_kw,
            "is_base_candidate": (
                row["pv_kw"] <= PV_IDLE_THRESHOLD_KW
                and order_kw <= EV_IDLE_THRESHOLD_KW
                and abs(row["ess_kw"]) <= ESS_IDLE_THRESHOLD_KW
            ),
        }
        night_rows.append(enriched)
        hourly_groups[ts.replace(minute=0, second=0, microsecond=0)].append(enriched)

    strict_hour_rows = []
    relaxed_hour_rows = []
    for hour_ts, rows in sorted(hourly_groups.items()):
        if len(rows) != 12:
            continue
        avg_load_kw = mean([row["total_load_kw"] for row in rows])
        avg_pv_kw = mean([row["pv_kw"] for row in rows])
        avg_order_kw = mean([row["order_kw"] for row in rows])
        avg_ess_kw = mean([row["ess_kw"] for row in rows])
        max_pv_kw = max(row["pv_kw"] for row in rows)
        max_order_kw = max(row["order_kw"] for row in rows)
        max_abs_ess_kw = max(abs(row["ess_kw"]) for row in rows)
        strict = all(row["is_base_candidate"] for row in rows)
        relaxed_count = sum(1 for row in rows if row["is_base_candidate"])
        relaxed = relaxed_count >= 10

        item = {
            "时间": hour_ts.strftime("%Y-%m-%d %H:00:00"),
            "日期": hour_ts.strftime("%Y%m%d"),
            "小时": f"{hour_ts.hour:02d}:00",
            "平均基础负荷(kW)": round(avg_load_kw, 4),
            "该小时电量(度)": round(avg_load_kw, 4),
            "平均光伏功率(kW)": round(avg_pv_kw, 4),
            "平均订单充电功率(kW)": round(avg_order_kw, 4),
            "平均储能功率(kW)": round(avg_ess_kw, 4),
            "最大光伏功率(kW)": round(max_pv_kw, 4),
            "最大订单充电功率(kW)": round(max_order_kw, 4),
            "最大储能绝对功率(kW)": round(max_abs_ess_kw, 4),
            "满足条件的5分钟点数": relaxed_count,
        }
        if strict:
            strict_hour_rows.append(item)
        if relaxed:
            relaxed_hour_rows.append(item)

    out_dir = ROOT / "报告" / "json"
    out_dir.mkdir(parents=True, exist_ok=True)
    strict_csv = out_dir / "夜间基础负荷_202606_严格样本.csv"
    relaxed_csv = out_dir / "夜间基础负荷_202606_宽松样本.csv"
    summary_csv = out_dir / "夜间基础负荷_202606_按小时统计.csv"
    report_md = ROOT / "报告" / "6月夜间基础负荷分析.md"

    fieldnames = [
        "时间",
        "日期",
        "小时",
        "平均基础负荷(kW)",
        "该小时电量(度)",
        "平均光伏功率(kW)",
        "平均订单充电功率(kW)",
        "平均储能功率(kW)",
        "最大光伏功率(kW)",
        "最大订单充电功率(kW)",
        "最大储能绝对功率(kW)",
        "满足条件的5分钟点数",
    ]

    def write_csv(path, rows):
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    write_csv(strict_csv, strict_hour_rows)
    write_csv(relaxed_csv, relaxed_hour_rows)

    by_hour = defaultdict(list)
    for row in strict_hour_rows:
        by_hour[row["小时"]].append(row["平均基础负荷(kW)"])
    with open(summary_csv, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["小时", "严格样本数", "平均基础负荷(kW)", "P10(kW)", "P50(kW)", "P90(kW)"])
        for hour in sorted(by_hour):
            vals = by_hour[hour]
            writer.writerow(
                [
                    hour,
                    len(vals),
                    round(mean(vals), 4),
                    round(percentile(vals, 0.1), 4),
                    round(percentile(vals, 0.5), 4),
                    round(percentile(vals, 0.9), 4),
                ]
            )

    strict_loads = [row["平均基础负荷(kW)"] for row in strict_hour_rows]
    relaxed_loads = [row["平均基础负荷(kW)"] for row in relaxed_hour_rows]
    top_strict = sorted(strict_hour_rows, key=lambda row: row["平均基础负荷(kW)"], reverse=True)[:10]
    low_strict = sorted(strict_hour_rows, key=lambda row: row["平均基础负荷(kW)"])[:10]

    lines = []
    lines.append("# 6月夜间基础负荷分析")
    lines.append("")
    lines.append("- 目标：按新规则识别真正的夜间基础负荷。")
    lines.append("- 夜间定义：`18:00-23:59` 与 `00:00-05:59`。")
    lines.append(
        f"- 严格筛选条件：`光伏功率 <= {PV_IDLE_THRESHOLD_KW}kW`、`订单充电功率 <= {EV_IDLE_THRESHOLD_KW}kW`、`|储能功率| <= {ESS_IDLE_THRESHOLD_KW}kW`，且整小时 12 个 5 分钟点全部满足。"
    )
    lines.append("- 宽松样本：整小时至少 `10/12` 个 5 分钟点满足上述条件，仅作为参考。")
    lines.append("")
    lines.append("## 样本结果")
    lines.append(f"- 夜间 5 分钟总样本数：`{len(night_rows)}`")
    lines.append(f"- 严格样本小时数：`{len(strict_hour_rows)}`")
    lines.append(f"- 宽松样本小时数：`{len(relaxed_hour_rows)}`")
    if strict_loads:
        lines.append(f"- 严格样本平均基础负荷：`{mean(strict_loads):.2f} kW`")
        lines.append(f"- 严格样本中位基础负荷：`{percentile(strict_loads, 0.5):.2f} kW`")
        lines.append(f"- 严格样本 P10-P90：`{percentile(strict_loads, 0.1):.2f} - {percentile(strict_loads, 0.9):.2f} kW`")
    if relaxed_loads:
        lines.append(f"- 宽松样本平均基础负荷：`{mean(relaxed_loads):.2f} kW`")
    lines.append("")
    lines.append("## 按小时统计")
    if by_hour:
        for hour in sorted(by_hour):
            vals = by_hour[hour]
            lines.append(
                f"- `{hour}`: 样本 `{len(vals)}` 小时，平均 `{mean(vals):.2f} kW`，P50 `{percentile(vals, 0.5):.2f} kW`"
            )
    else:
        lines.append("- 严格条件下未找到有效整小时样本。")
    lines.append("")
    lines.append("## 严格样本最高负荷 Top 10")
    if top_strict:
        for row in top_strict:
            lines.append(
                f"- `{row['时间']}`: 基础负荷 `{row['平均基础负荷(kW)']:.2f} kW`，订单 `{row['平均订单充电功率(kW)']:.2f} kW`，储能 `{row['平均储能功率(kW)']:.2f} kW`"
            )
    else:
        lines.append("- 无")
    lines.append("")
    lines.append("## 严格样本最低负荷 Top 10")
    if low_strict:
        for row in low_strict:
            lines.append(
                f"- `{row['时间']}`: 基础负荷 `{row['平均基础负荷(kW)']:.2f} kW`，订单 `{row['平均订单充电功率(kW)']:.2f} kW`，储能 `{row['平均储能功率(kW)']:.2f} kW`"
            )
    else:
        lines.append("- 无")
    lines.append("")
    lines.append("## 说明")
    lines.append("- 该结果更接近 `宿舍生活用电 + 设备空载/待机负荷 + 站内基础辅电`。")
    lines.append("- 该结果不包含主动储能充放电，也尽量排除了车辆充电与夜间无光伏之外的影响。")
    lines.append("- 若严格样本偏少，说明 6 月夜间大部分时间系统并不处于完全静止状态。")

    report_md.write_text("\n".join(lines), encoding="utf-8")

    print("strict_hours", len(strict_hour_rows))
    print("relaxed_hours", len(relaxed_hour_rows))
    if strict_loads:
        print("strict_mean_kw", round(mean(strict_loads), 4))
        print("strict_median_kw", round(percentile(strict_loads, 0.5), 4))
        print("strict_p10_kw", round(percentile(strict_loads, 0.1), 4))
        print("strict_p90_kw", round(percentile(strict_loads, 0.9), 4))
    print("strict_csv", strict_csv)
    print("relaxed_csv", relaxed_csv)
    print("summary_csv", summary_csv)
    print("report_md", report_md)


if __name__ == "__main__":
    main()
