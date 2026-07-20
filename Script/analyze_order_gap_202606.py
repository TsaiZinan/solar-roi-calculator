from __future__ import annotations

import csv
import glob
import json
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/cai/SynologyDrive/Project/#ProjectWork-000000-光伏收益计算")
ORDER_GLOB = str(ROOT / "数据" / "TEMP" / "*.xlsx")
DAILY_JSON_GLOB = str(ROOT / "报告" / "json" / "每日收益分析_202606*.json")


def parse_dt(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def to_float(value):
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def load_model_day_hour():
    model = {}
    for path in sorted(glob.glob(DAILY_JSON_GLOB)):
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        date_str = payload.get("date") or Path(path).stem[-8:]
        for row in payload.get("hourly_stats", []):
            hour = int(str(row["hour"]).split(":")[0])
            model[(date_str, hour)] = float(row.get("charging_pile_load_kwh", 0.0))
    return model


def load_actual_day_hour():
    order_path = Path(sorted(glob.glob(ORDER_GLOB))[0])
    wb = load_workbook(order_path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {name: i for i, name in enumerate(header)}

    actual = defaultdict(float)
    valid_orders = 0
    fallback_orders = 0
    channel_stats = defaultdict(lambda: {"orders": 0, "kwh": 0.0})

    for row in rows:
        status = row[idx["充电状态"]]
        pay_status = row[idx["支付状态"]]
        abnormal = row[idx["是否异常"]]
        if not (status == "已完成" and pay_status == "已支付" and abnormal in ("正常", None, "")):
            continue

        start = parse_dt(row[idx["开始充电时间"]])
        end = parse_dt(row[idx["充电完成时间"]])
        kwh = to_float(row[idx["充电量（度）"]])
        source = str(row[idx["订单来源"]] or "NA")
        if not start or kwh <= 0:
            continue

        valid_orders += 1
        channel_stats[source]["orders"] += 1
        channel_stats[source]["kwh"] += kwh

        if not end or end <= start:
            actual[(start.strftime("%Y%m%d"), start.hour)] += kwh
            fallback_orders += 1
            continue

        total_seconds = (end - start).total_seconds()
        if total_seconds <= 0:
            actual[(start.strftime("%Y%m%d"), start.hour)] += kwh
            fallback_orders += 1
            continue

        cursor = start
        while cursor < end:
            hour_start = cursor.replace(minute=0, second=0, microsecond=0)
            next_hour = hour_start + timedelta(hours=1)
            segment_end = min(end, next_hour)
            overlap_seconds = (segment_end - cursor).total_seconds()
            if overlap_seconds > 0:
                share = kwh * overlap_seconds / total_seconds
                actual[(hour_start.strftime("%Y%m%d"), hour_start.hour)] += share
            cursor = segment_end

    return order_path, actual, valid_orders, fallback_orders, channel_stats


def write_csv(path, header, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def main():
    model = load_model_day_hour()
    order_path, actual, valid_orders, fallback_orders, channel_stats = load_actual_day_hour()

    all_keys = sorted(set(model) | set(actual))
    by_day = defaultdict(lambda: {"model": 0.0, "actual": 0.0, "diff": 0.0})
    by_hour = defaultdict(lambda: {"model": 0.0, "actual": 0.0, "diff": 0.0})
    detail_rows = []

    for date_str, hour in all_keys:
        model_kwh = model.get((date_str, hour), 0.0)
        actual_kwh = actual.get((date_str, hour), 0.0)
        diff_kwh = model_kwh - actual_kwh
        detail_rows.append(
            [
                date_str,
                f"{hour:02d}:00",
                round(model_kwh, 4),
                round(actual_kwh, 4),
                round(diff_kwh, 4),
            ]
        )
        by_day[date_str]["model"] += model_kwh
        by_day[date_str]["actual"] += actual_kwh
        by_day[date_str]["diff"] += diff_kwh
        by_hour[hour]["model"] += model_kwh
        by_hour[hour]["actual"] += actual_kwh
        by_hour[hour]["diff"] += diff_kwh

    out_json_dir = ROOT / "报告" / "json"
    out_json_dir.mkdir(parents=True, exist_ok=True)
    detail_csv = out_json_dir / "充电量差异_202606_逐日逐小时.csv"
    day_csv = out_json_dir / "充电量差异_202606_逐日汇总.csv"
    hour_csv = out_json_dir / "充电量差异_202606_逐小时汇总.csv"
    summary_md = ROOT / "报告" / "6月充电量差异分布分析.md"

    write_csv(
        detail_csv,
        ["日期", "小时", "模型充电量(度)", "订单充电量(度)", "差异(模型-订单)(度)"],
        detail_rows,
    )

    day_rows = []
    for date_str in sorted(by_day):
        model_kwh = by_day[date_str]["model"]
        actual_kwh = by_day[date_str]["actual"]
        diff_kwh = by_day[date_str]["diff"]
        ratio = diff_kwh / actual_kwh if actual_kwh else ""
        day_rows.append(
            [
                date_str,
                round(model_kwh, 4),
                round(actual_kwh, 4),
                round(diff_kwh, 4),
                round(ratio, 6) if ratio != "" else "",
            ]
        )
    write_csv(day_csv, ["日期", "模型充电量(度)", "订单充电量(度)", "差异(度)", "差异占订单比"], day_rows)

    hour_rows = []
    for hour in range(24):
        model_kwh = by_hour[hour]["model"]
        actual_kwh = by_hour[hour]["actual"]
        diff_kwh = by_hour[hour]["diff"]
        ratio = diff_kwh / actual_kwh if actual_kwh else ""
        hour_rows.append(
            [
                f"{hour:02d}:00",
                round(model_kwh, 4),
                round(actual_kwh, 4),
                round(diff_kwh, 4),
                round(ratio, 6) if ratio != "" else "",
            ]
        )
    write_csv(hour_csv, ["小时", "模型充电量(度)", "订单充电量(度)", "差异(度)", "差异占订单比"], hour_rows)

    model_total = sum(item["model"] for item in by_day.values())
    actual_total = sum(item["actual"] for item in by_day.values())
    diff_total = sum(item["diff"] for item in by_day.values())

    top_days = sorted(by_day.items(), key=lambda kv: kv[1]["diff"], reverse=True)[:10]
    top_hours = sorted(by_hour.items(), key=lambda kv: kv[1]["diff"], reverse=True)
    top_day_hours = sorted(detail_rows, key=lambda row: row[4], reverse=True)[:20]
    low_day_hours = sorted(detail_rows, key=lambda row: row[4])[:20]

    lines = []
    lines.append("# 6月充电量差异分布分析")
    lines.append("")
    lines.append(f"- 订单文件: `{order_path.name}`")
    lines.append(f"- 有效订单数: `{valid_orders}`")
    lines.append(f"- 回退到起始小时分配的订单数: `{fallback_orders}`")
    lines.append(f"- 模型充电量合计: `{model_total:.2f}` 度")
    lines.append(f"- 订单充电量合计: `{actual_total:.2f}` 度")
    lines.append(f"- 差异合计: `{diff_total:.2f}` 度")
    lines.append(f"- 差异占订单比: `{(diff_total / actual_total * 100) if actual_total else 0:.2f}%`")
    lines.append("")
    lines.append("## 逐日差异 Top 10")
    for date_str, data in top_days:
        ratio = data["diff"] / data["actual"] * 100 if data["actual"] else 0.0
        lines.append(
            f"- `{date_str}`: 模型 `{data['model']:.2f}`，订单 `{data['actual']:.2f}`，差异 `{data['diff']:.2f}`，占订单 `{ratio:.1f}%`"
        )
    lines.append("")
    lines.append("## 逐小时差异汇总")
    for hour, data in top_hours:
        ratio = data["diff"] / data["actual"] * 100 if data["actual"] else 0.0
        lines.append(
            f"- `{hour:02d}:00`: 模型 `{data['model']:.2f}`，订单 `{data['actual']:.2f}`，差异 `{data['diff']:.2f}`，占订单 `{ratio:.1f}%`"
        )
    lines.append("")
    lines.append("## 单日单小时差异 Top 20")
    for date_str, hour_label, model_kwh, actual_kwh, diff_kwh in top_day_hours:
        lines.append(
            f"- `{date_str} {hour_label}`: 模型 `{model_kwh:.2f}`，订单 `{actual_kwh:.2f}`，差异 `{diff_kwh:.2f}`"
        )
    lines.append("")
    lines.append("## 单日单小时低估 Top 20")
    for date_str, hour_label, model_kwh, actual_kwh, diff_kwh in low_day_hours:
        lines.append(
            f"- `{date_str} {hour_label}`: 模型 `{model_kwh:.2f}`，订单 `{actual_kwh:.2f}`，差异 `{diff_kwh:.2f}`"
        )

    summary_md.write_text("\n".join(lines), encoding="utf-8")

    print("detail_csv", detail_csv)
    print("day_csv", day_csv)
    print("hour_csv", hour_csv)
    print("summary_md", summary_md)
    print("model_total", round(model_total, 2))
    print("actual_total", round(actual_total, 2))
    print("diff_total", round(diff_total, 2))
    print("top_days")
    for date_str, data in top_days[:5]:
        print(date_str, round(data["diff"], 2), round(data["model"], 2), round(data["actual"], 2))
    print("top_hours")
    for hour, data in top_hours[:8]:
        print(f"{hour:02d}:00", round(data["diff"], 2), round(data["model"], 2), round(data["actual"], 2))
    print("top_channels")
    for source, data in sorted(channel_stats.items(), key=lambda kv: kv[1]["kwh"], reverse=True)[:8]:
        print(source, data["orders"], round(data["kwh"], 2))


if __name__ == "__main__":
    main()
